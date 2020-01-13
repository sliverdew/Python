#/bin/sh
#
import os
import shutil
import openpyxl

def iter_files(rootDir,file_list,dir_list):
	#遍历根目录
	for root,dirs,files in os.walk(rootDir):
		for file in files:
			file_name = os.path.join(root,file)
			file_list.append(file_name)
			dir_list.append(root)
			#print(file_name)
#原始目录路径
rootDir = r"K:\FaceCapture"
#用来存放所有的文件路径
file_list = []
#用来存放所有的目录路径
dir_list = []		
q = 0
iter_files(rootDir,file_list,dir_list)
outwb = Workbook()#打开一个将写的文件
outws = outwb.create_sheet(title="test") #在将写的文件创建sheet
for q in range(len(dir_list)):
	outws.cell(row = 1 , column = q+1).value = dir_list[q]
outwb.save('test.xls')
